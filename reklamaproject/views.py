from rest_framework import viewsets, permissions, filters
from rest_framework.views import APIView
from .models import MetroLine, Station, Position, Advertisement, AdvertisementArchive, Ijarachi, Turi,TarkibShartnomaSummasi,TarkibAdvertisementArchiveShartnomaSummasi, ShartnomaSummasi, ShartnomaSummasiArchive, Depo, HarakatTarkibi, TarkibPosition, TarkibAdvertisement, TarkibAdvertisementArchive
from .serializers import (
    MetroLineSerializer, StationSerializer,
    PositionSerializer, AdvertisementSerializer, AdvertisementArchiveSerializer, 
    CreateAdvertisementSerializer, ExportAdvertisementSerializer,AdvertisementStatisticsSerializer,TarkibAdvertisementArchiveShartnomaSummasiSerializer,
    IjarachiSerializers, TuriSerializer, ShartnomaSummasiSerializer, UpdateAdvertisementSerializer,TarkibShartnomaSummasiSerializer,
     TarkibAdvertisementSerializer, TarkibAdvertisementArchiveSerializer,CreateTarkibAdvertisementSerializer,UpdateTarkibAdvertisementSerializer,TarkibPositionSerializer, DepoSerializer, HarakatTarkibiSerializer
)
from .pagination import CustomPagination
from rest_framework import status
from django.db import transaction
from rest_framework.decorators import api_view, permission_classes, action
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response
from django_filters.rest_framework import DjangoFilterBackend
from openpyxl.drawing.image import Image as XLImage
from openpyxl.utils import get_column_letter
import openpyxl
from django.http import HttpResponse
from datetime import date, timedelta
from django.db.models import Q
from openpyxl import Workbook
from rest_framework.renderers import BaseRenderer
from drf_spectacular.utils import extend_schema, OpenApiParameter
from rest_framework.filters import SearchFilter
from django.utils import timezone
import cv2
from reportlab.pdfgen import canvas
import pytesseract
pytesseract.pytesseract.tesseract_cmd = r"C:\Program Files\Tesseract-OCR\tesseract.exe"
from django.conf import settings
from drf_spectacular.utils import extend_schema, OpenApiParameter
from reportlab.lib.pagesizes import A4, landscape
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer,HRFlowable, Image, PageBreak
from reportlab.graphics.shapes import Drawing, String
from reportlab.graphics.charts.barcharts import HorizontalBarChart, VerticalBarChart
from reportlab.lib import colors
import io
from reportlab.lib.styles import getSampleStyleSheet
import requests
from django.utils.timezone import now
from django.db.models import Count, Sum
from io import BytesIO
import os
from reportlab.lib.units import inch
from django.http import FileResponse
from datetime import datetime
from reportlab.lib.styles import ParagraphStyle
from decimal import Decimal
from reportlab.lib.enums import TA_LEFT, TA_CENTER, TA_RIGHT, TA_JUSTIFY
class XLSXRenderer(BaseRenderer):
    media_type = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    format = 'xlsx'
    charset = None

    def render(self, data, media_type=None, renderer_context=None):
        return data



def generate_pdf_detail(filename, title, data_list):
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=A4,
        title=title
    )

    styles = getSampleStyleSheet()
    elements = []

    # Title
    elements.append(Paragraph(f"<b>{title}</b>", styles["Title"]))
    elements.append(Spacer(1, 12))

    # Har bir reklama uchun alohida blok
    for idx, item in enumerate(data_list, start=1):
        elements.append(Paragraph(f"<b>Reklama #{idx}</b>", styles["Heading2"]))
        elements.append(Spacer(1, 6))

        # Agar image bo‘lsa – tepada chiqaramiz
        if "image" in item and item["image"]:
            try:
                img_resp = requests.get(item["image"], timeout=5)
                if img_resp.status_code == 200:
                    img_data = io.BytesIO(img_resp.content)
                    img = Image(img_data, width=150, height=100)
                    elements.append(img)
                    elements.append(Spacer(1, 6))
            except Exception:
                elements.append(Paragraph("<b>Rasm yuklanmadi</b>", styles["Normal"]))
                elements.append(Spacer(1, 6))

        # Qolgan ma’lumotlar
        for key, value in item.items():
            if key == "image":  
                continue  # image allaqachon chiqarildi
            text = f'<font color="blue"><b>{key}:</b></font> {value}'
            elements.append(Paragraph(text, styles["Normal"]))
            elements.append(Spacer(1, 3))

        # Ajratish chizig‘i
        elements.append(Spacer(1, 6))
        elements.append(HRFlowable(width="100%", color=colors.grey, thickness=0.7, lineCap='round'))
        elements.append(Spacer(1, 12))

    doc.build(elements)

    buffer.seek(0)
    response = HttpResponse(buffer, content_type="application/pdf")
    response["Content-Disposition"] = f'attachment; filename="{filename}.pdf"'
    return response

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def get_me(request):
    return Response({
        "id": request.user.id,
        "username": request.user.username,
        "is_superuser": request.user.is_superuser,
        "email": request.user.email,
    })

class AuthenticatedCRUDPermission(permissions.BasePermission):
    def has_permission(self, request, view):
        return request.user and request.user.is_authenticated

class MetroLineViewSet(viewsets.ModelViewSet):
    queryset = MetroLine.objects.all()
    serializer_class = MetroLineSerializer
    permission_classes = [AuthenticatedCRUDPermission]
    pagination_class = CustomPagination

class StationViewSet(viewsets.ModelViewSet):
    queryset = Station.objects.all()
    serializer_class = StationSerializer
    permission_classes = [AuthenticatedCRUDPermission]
    filter_backends = [DjangoFilterBackend]
    filterset_fields = ['line']
    # pagination_class = CustomPagination


class Stationimage(APIView):
    permission_classes = [IsAuthenticated]

    def put(self, request, pk):
        try:
            station = Station.objects.get(pk=pk)
        except Station.DoesNotExist:
            return Response({"error": "Bunday bekat topilmadi"}, status=404)

        image = request.FILES.get("schema_image")
        if not image:
            return Response({"error": "Rasm yuborilmadi"}, status=400)

        station.schema_image = image
        station.save()
        return Response({"message": "Rasm yangilandi", "id": station.id})



class PositionViewSet(viewsets.ModelViewSet):
    queryset = Position.objects.select_related('station').prefetch_related('advertisement').all().order_by('-created_at')
    serializer_class = PositionSerializer
    permission_classes = [AuthenticatedCRUDPermission]
    filter_backends = [DjangoFilterBackend, filters.OrderingFilter, filters.SearchFilter]
    filterset_fields = ['station']
    search_fields = ['number']
    pagination_class = CustomPagination

    def perform_create(self, serializer):
        serializer.save(
            created_by=self.request.user,
            created_at=timezone.now()
        )

    def perform_destroy(self, instance):
        """
        Position o‘chirilsa -> unga biriktirilgan Advertisement bo‘lsa,
        avval arxivga ko‘chadi keyin o‘chadi, so‘ng Position o‘chiriladi.
        """
        ad = getattr(instance, 'advertisement', None)
        if ad:
            station = instance.station
            line = station.line if station else None

            AdvertisementArchive.objects.create(
                original_ad=ad,
                user=self.request.user,
                position=instance,
                line=line,
                station=station,
                Reklama_nomi=ad.Reklama_nomi,
                Qurilma_turi=ad.Qurilma_turi,
                Ijarachi=ad.Ijarachi,
                Shartnoma_raqami=ad.Shartnoma_raqami,
                Shartnoma_muddati_boshlanishi=ad.Shartnoma_muddati_boshlanishi,
                Shartnoma_tugashi=ad.Shartnoma_tugashi,
                O_lchov_birligi=ad.O_lchov_birligi,
                Qurilma_narxi=ad.Qurilma_narxi,
                Egallagan_maydon=ad.Egallagan_maydon,
                Shartnoma_summasi=ad.Shartnoma_summasi,
                Shartnoma_fayl=ad.Shartnoma_fayl,
                photo=ad.photo,
                contact_number=ad.contact_number,
            )
            ad.delete()

        instance.delete()




class IjarachiViewSet(viewsets.ModelViewSet):
    queryset = Ijarachi.objects.all().order_by("-id")
    serializer_class = IjarachiSerializers
    permission_classes = [AuthenticatedCRUDPermission]  
    filter_backends = [filters.SearchFilter, filters.OrderingFilter, DjangoFilterBackend]
    search_fields = ["name", "contact_number"]
    ordering_fields = ["name", "id"]
    filterset_fields = ["name"]
    pagination_class = CustomPagination

    def perform_create(self, serializer):
        serializer.save()

    def perform_update(self, serializer):
        serializer.save()

    # ======================
    # Excel eksport
    # ======================
    @action(detail=False, methods=['get'], url_path='export-excel')
    def export_excel(self, request):
        queryset = self.filter_queryset(self.get_queryset())

        wb = Workbook()
        ws = wb.active
        ws.title = "Ijarachilar"

        headers = ["Rasm", "Ism", "Kontakt raqam"]
        ws.append(headers)

        for ijarachi in queryset:
            row = [
                "",  # Rasm uchun bo‘sh joy qoldiramiz, keyin quyida joylashtiramiz
                ijarachi.name,
                ijarachi.contact_number,
            ]
            ws.append(row)

            # Agar rasm mavjud bo‘lsa, Excelga joylashtirish
            if hasattr(ijarachi, 'photo') and ijarachi.photo:
                try:
                    img_data = requests.get(ijarachi.photo.url).content
                    image = XLImage(BytesIO(img_data))
                    # Rasmni shu qator, birinchi ustunga joylashtiramiz
                    image.anchor = f"A{ws.max_row}"
                    ws.add_image(image)
                except Exception as e:
                    print(f"Rasm qo'shishda xatolik: {e}")

        # ustun kengligi
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 20
        ws.row_dimensions[1].height = 30  # header uchun

        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        response['Content-Disposition'] = 'attachment; filename=ijarachilar.xlsx'
        wb.save(response)
        return response

    # ======================
    # PDF eksport
    # ======================
    @action(detail=False, methods=['get'], url_path='export-pdf')
    def export_pdf(self, request):
        queryset = self.filter_queryset(self.get_queryset())

        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4)
        elements = []

        styles = getSampleStyleSheet()
        title_style = styles['Heading1']
        normal_style = styles['Normal']

        # Title
        elements.append(Paragraph("Ijarachilar Ro'yxati", title_style))
        elements.append(Spacer(1, 20))

        # Jadval sarlavhalari
        data = [["ID", "Logo", "Ism", "Kontakt raqam"]]

        for ijarachi in queryset:
            # Logo rasmi
            if ijarachi.logo and os.path.exists(ijarachi.logo.path):
                try:
                    img = Image(ijarachi.logo.path, width=40, height=40)
                except Exception:
                    img = ""
            else:
                img = ""

            row = [
                str(ijarachi.id),
                img,
                ijarachi.name,
                ijarachi.contact_number or ""
            ]
            data.append(row)

        # Jadval
        table = Table(data, colWidths=[40, 50, 200, 120])
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 12),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.black),
        ]))

        elements.append(table)
        doc.build(elements)

        buffer.seek(0)
        response = HttpResponse(buffer, content_type='application/pdf')
        response['Content-Disposition'] = 'attachment; filename="ijarachilar.pdf"'
        return response




def generate_pdf_detail(filename, title, data_list):
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4)
    elements = []
    styles = getSampleStyleSheet()

    # Report title
    elements.append(Paragraph(title, styles['Title']))
    elements.append(Spacer(1, 12))

    for idx, item in enumerate(data_list, start=1):
        elements.append(Paragraph(f"Reklama #{idx}", styles['Heading2']))
        elements.append(Spacer(1, 6))

        # Rasm
        if item.get("image") and os.path.exists(item["image"].replace("/media/", "media/")):
            try:
                img_path = item["image"].replace("/media/", "media/")
                elements.append(Image(img_path, width=3*inch, height=2*inch))
                elements.append(Spacer(1, 12))
            except Exception:
                elements.append(Paragraph("Rasm yuklanmadi", styles['Normal']))
        else:
            elements.append(Paragraph("Rasm yuklanmadi", styles['Normal']))

        # Table
        table_data = []
        for key, value in item.items():
            if key == "image":
                continue
            table_data.append([key, str(value)])

        table = Table(table_data, colWidths=[150, 300])
        table.setStyle(TableStyle([
            ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
            ("BACKGROUND", (0, 0), (-1, 0), colors.lightgrey),
            ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ]))
        elements.append(table)
        elements.append(Spacer(1, 20))

    doc.build(elements)
    buffer.seek(0)

    response = HttpResponse(buffer, content_type="application/pdf")
    response["Content-Disposition"] = f'attachment; filename="{filename}.pdf"'
    return response






class TuriViewSet(viewsets.ModelViewSet):
    queryset = Turi.objects.all().order_by("-id")
    serializer_class = TuriSerializer
    search_fields = ['qurilmaturi']
    permission_classes = [AuthenticatedCRUDPermission]
    filter_backends = [filters.SearchFilter, filters.OrderingFilter, DjangoFilterBackend]
    ordering_fields = ['qurilmaturi', 'id']
    pagination_class = CustomPagination



class AdvertisementViewSet(viewsets.ModelViewSet):
    queryset = Advertisement.objects.all()
    serializer_class = AdvertisementSerializer
    permission_classes = [AuthenticatedCRUDPermission]
    filter_backends = [filters.SearchFilter, filters.OrderingFilter, DjangoFilterBackend]
    search_fields = ['Reklama_nomi', 'Shartnoma_raqami']
    ordering_fields = ['created_at', 'Qurilma_narxi']
    filterset_fields = ['position__station', 'position__station__line']
    # pagination_class = CustomPagination

    def get_serializer_class(self):
        if self.action == 'create':
            return CreateAdvertisementSerializer
        elif self.action in ['update', 'partial_update']:
            return UpdateAdvertisementSerializer
        return AdvertisementSerializer


    def perform_create(self, serializer):
        serializer.save()

    
    def perform_update(self, serializer):
        old_instance = serializer.instance  # eski obyekt
        print("PERFORM_UPDATE - OLD DATA:", {
            'Reklama_nomi': old_instance.Reklama_nomi,
            'Ijarachi': old_instance.Ijarachi
        })

        # --- ARXIVGA ESKISI SAQLANADI ---
        archive = AdvertisementArchive.objects.create(
            original_ad=old_instance,
            user=self.request.user,
            position=old_instance.position,
            line=old_instance.position.station.line if old_instance.position and old_instance.position.station else None,
            station=old_instance.position.station if old_instance.position else None,
            Reklama_nomi=old_instance.Reklama_nomi,
            Qurilma_turi=old_instance.Qurilma_turi,
            Ijarachi=old_instance.Ijarachi,
            Shartnoma_raqami=old_instance.Shartnoma_raqami,
            Shartnoma_muddati_boshlanishi=old_instance.Shartnoma_muddati_boshlanishi,
            Shartnoma_tugashi=old_instance.Shartnoma_tugashi,
            O_lchov_birligi=old_instance.O_lchov_birligi,
            Qurilma_narxi=old_instance.Qurilma_narxi,
            Egallagan_maydon=old_instance.Egallagan_maydon,
            Shartnoma_summasi=old_instance.Shartnoma_summasi,
            Shartnoma_fayl=old_instance.Shartnoma_fayl,
            photo=old_instance.photo,
        )

        # Tolovlarni arxivga
        for t in old_instance.tolovlar.all():
            ShartnomaSummasiArchive.objects.create(
                archive=archive,
                Shartnomasummasi=t.Shartnomasummasi,
                created_at=t.created_at
            )

        # --- DB GA YANGILASH ---
        updated_instance = serializer.save()  # bu holda DB yangi qiymat bilan yangilanadi
        print("PERFORM_UPDATE - UPDATED DATA:", {
            'Reklama_nomi': updated_instance.Reklama_nomi,
            'Ijarachi': updated_instance.Ijarachi
        })

        return updated_instance




        

    # --- OVERRIDE UPDATE METHOD ---
    def update(self, request, *args, **kwargs):
        partial = kwargs.pop('partial', False)
        instance = self.get_object()
        serializer = self.get_serializer(instance, data=request.data, partial=partial)
        serializer.is_valid(raise_exception=True)
        updated_instance = self.perform_update(serializer)
        
        # --- Responseda yangi obyektni serializer orqali qaytarish ---
        return Response(self.get_serializer(updated_instance).data)












    def perform_destroy(self, instance):
        """Delete tugmasi bosilganda reklama arxivga ko‘chadi va keyin o‘chadi"""
        station = instance.position.station if instance.position else None
        line = station.line if station else None
        old_instance = self.get_object()

        archive = AdvertisementArchive.objects.create(
            original_ad=instance,
            user=self.request.user,
            position=instance.position,
            line=line,
            station=station,
            Reklama_nomi=instance.Reklama_nomi,
            Qurilma_turi=instance.Qurilma_turi,
            Ijarachi=instance.Ijarachi,
            Shartnoma_raqami=instance.Shartnoma_raqami,
            Shartnoma_muddati_boshlanishi=instance.Shartnoma_muddati_boshlanishi,
            Shartnoma_tugashi=instance.Shartnoma_tugashi,
            O_lchov_birligi=instance.O_lchov_birligi,
            Qurilma_narxi=instance.Qurilma_narxi,
            Egallagan_maydon=instance.Egallagan_maydon,
            Shartnoma_summasi=instance.Shartnoma_summasi,
            Shartnoma_fayl=instance.Shartnoma_fayl,
            photo=instance.photo,

        )
        for tolov in old_instance.tolovlar.all():
            ShartnomaSummasiArchive.objects.create(
                archive=archive,
                Shartnomasummasi=tolov.Shartnomasummasi,
                created_at=tolov.created_at
            )
        instance.delete()

    @action(detail=True, methods=['get', 'post'], url_path='export')
    def export_advertisement(self, request, pk=None):
        source_ad = self.get_object()

        if request.method == 'GET':
            serializer = ExportAdvertisementSerializer()
            return Response(serializer.data)

        serializer = ExportAdvertisementSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        target_position = serializer.validated_data['position']

        with transaction.atomic():
            # 1. Target joydagi reklamani arxivga yuboramiz (agar mavjud bo‘lsa)
            target_ad = getattr(target_position, 'advertisement', None)
            if target_ad:
                target_station = target_ad.position.station
                target_line = target_station.line

                AdvertisementArchive.objects.create(
                    original_ad=target_ad,
                    user=self.request.user,
                    position=target_ad.position,
                    line=target_line,
                    station=target_station,
                    Reklama_nomi=target_ad.Reklama_nomi,
                    Qurilma_turi=target_ad.Qurilma_turi,
                    Ijarachi=target_ad.Ijarachi,
                    Shartnoma_raqami=target_ad.Shartnoma_raqami,
                    Shartnoma_muddati_boshlanishi=target_ad.Shartnoma_muddati_boshlanishi,
                    Shartnoma_tugashi=target_ad.Shartnoma_tugashi,
                    O_lchov_birligi=target_ad.O_lchov_birligi,
                    Qurilma_narxi=target_ad.Qurilma_narxi,
                    Egallagan_maydon=target_ad.Egallagan_maydon,
                    Shartnoma_summasi=target_ad.Shartnoma_summasi,
                    Shartnoma_fayl=target_ad.Shartnoma_fayl,
                    photo=target_ad.photo,
                    contact_number=source_ad.Ijarachi.contact_number if source_ad.Ijarachi else None

                )
                # 2. Target joydagi eski reklamani o‘chiramiz
                target_ad.delete()

            # 3. Source reklamani target pozitsiyaga nusxalaymiz
            Advertisement.objects.create(
                user=self.request.user,
                position=target_position,
                Reklama_nomi=source_ad.Reklama_nomi,
                Qurilma_turi=source_ad.Qurilma_turi,
                Ijarachi=source_ad.Ijarachi,
                Shartnoma_raqami=source_ad.Shartnoma_raqami,
                Shartnoma_muddati_boshlanishi=source_ad.Shartnoma_muddati_boshlanishi,
                Shartnoma_tugashi=source_ad.Shartnoma_tugashi,
                O_lchov_birligi=source_ad.O_lchov_birligi,
                Qurilma_narxi=source_ad.Qurilma_narxi,
                Egallagan_maydon=source_ad.Egallagan_maydon,
                Shartnoma_summasi=source_ad.Shartnoma_summasi,
                Shartnoma_fayl=source_ad.Shartnoma_fayl,
                photo=source_ad.photo,
                contact_number=source_ad.Ijarachi.contact_number if source_ad.Ijarachi else None

            )

            source_station = source_ad.position.station
            source_line = source_station.line

            AdvertisementArchive.objects.create(
                original_ad=source_ad,
                user=self.request.user,
                position=source_ad.position,
                line=source_line,
                station=source_station,
                Reklama_nomi=source_ad.Reklama_nomi,
                Qurilma_turi=source_ad.Qurilma_turi,
                Ijarachi=source_ad.Ijarachi,
                Shartnoma_raqami=source_ad.Shartnoma_raqami,
                Shartnoma_muddati_boshlanishi=source_ad.Shartnoma_muddati_boshlanishi,
                Shartnoma_tugashi=source_ad.Shartnoma_tugashi,
                O_lchov_birligi=source_ad.O_lchov_birligi,
                Qurilma_narxi=source_ad.Qurilma_narxi,
                Egallagan_maydon=source_ad.Egallagan_maydon,
                Shartnoma_summasi=source_ad.Shartnoma_summasi,
                Shartnoma_fayl=source_ad.Shartnoma_fayl,
                photo=source_ad.photo,
                contact_number=source_ad.Ijarachi.contact_number if source_ad.Ijarachi else None

            )

            # 5. Source joyni bo‘shatamiz
            source_ad.delete()

        return Response({'detail': 'Reklama muvaffaqiyatli ko‘chirildi, arxivlandi va source joy tozalandi.'}, status=200)
    @action(detail=False, methods=['get'], url_path='export-excel')
    def export_excel(self, request):
        queryset = self.filter_queryset(self.get_queryset())

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Advertisements"

        headers = [
            "Reklama nomi", "Qurilma turi", "Ijarachi",
            "Shartnoma raqami", "Shartnoma boshlanishi", "Shartnoma tugashi",
            "O'lchov birligi", "Qurilma narxi", "Egallagan maydon", "Shartnoma summasi",
            "Position", "Station", "Line", "Contact number", "Created at", "Created by"
        ]
        ws.append(headers)

        for ad in queryset:
            row = [
                ad.Reklama_nomi,
                str(ad.Qurilma_turi),
                str(ad.Ijarachi) if ad.Ijarachi else "",
                ad.Shartnoma_raqami,
                ad.Shartnoma_muddati_boshlanishi.strftime("%Y-%m-%d") if ad.Shartnoma_muddati_boshlanishi else "",
                ad.Shartnoma_tugashi.strftime("%Y-%m-%d") if ad.Shartnoma_tugashi else "",
                ad.O_lchov_birligi,
                float(ad.Qurilma_narxi) if ad.Qurilma_narxi else 0,
                ad.Egallagan_maydon,
                float(ad.Shartnoma_summasi) if ad.Shartnoma_summasi else 0,
                ad.position.number if ad.position else "",
                ad.position.station.name if ad.position and ad.position.station else "",
                ad.position.station.line.name if ad.position and ad.position.station and ad.position.station.line else "",
                ad.Ijarachi.contact_number if ad.Ijarachi and hasattr(ad.Ijarachi, "contact_number") else "",
                ad.created_at.strftime("%Y-%m-%d %H:%M:%S") if ad.created_at else "",
                ad.user.username if ad.user else "",
            ]
            ws.append(row)

        # Ustunlarni avtomatik kengaytirish
        for i, column in enumerate(ws.columns, start=1):
            max_length = max((len(str(cell.value)) for cell in column if cell.value), default=0)
            ws.column_dimensions[get_column_letter(i)].width = max_length + 2

        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        response['Content-Disposition'] = 'attachment; filename=advertisements.xlsx'
        wb.save(response)
        return response

    @action(detail=False, methods=['get'], url_path='export-pdf')
    def export_pdf(self, request):
        queryset = self.filter_queryset(self.get_queryset())

        data_list = []
        for ad in queryset:
            data_list.append({
                "image": ad.photo.url if ad.photo else "",
                "Reklama nomi": ad.Reklama_nomi,
                "Qurilma turi": ad.Qurilma_turi,
                "Ijarachi": str(ad.Ijarachi) if ad.Ijarachi else "",
                "Shartnoma raqami": ad.Shartnoma_raqami,
                "Shartnoma boshlanishi": ad.Shartnoma_muddati_boshlanishi.strftime("%Y-%m-%d") if ad.Shartnoma_muddati_boshlanishi else "",
                "Shartnoma tugashi": ad.Shartnoma_tugashi.strftime("%Y-%m-%d") if ad.Shartnoma_tugashi else "",
                "O'lchov birligi": ad.O_lchov_birligi,
                "Qurilma narxi": float(ad.Qurilma_narxi) if ad.Qurilma_narxi else 0,
                "Egallagan maydon": ad.Egallagan_maydon,
                "Shartnoma summasi": float(ad.Shartnoma_summasi) if ad.Shartnoma_summasi else 0,
                "Joyi": ad.position.number if ad.position else "",
                "Bekati": ad.position.station.name if ad.position and ad.position.station else "",
                "Liniyasi": ad.position.station.line.name if ad.position and ad.position.station and ad.position.station.line else "",
                "Bog'lanish raqami": ad.Ijarachi.contact_number if ad.Ijarachi and hasattr(ad.Ijarachi, "contact_number") else "",
                "Yaratilgan vaqti": ad.created_at.strftime("%Y-%m-%d %H:%M:%S") if ad.created_at else "",
                "Kim tomonidan": ad.user.username if ad.user else "",
            })

        return generate_pdf_detail("advertisements", "Advertisements Report", data_list)



class AdvertisementArchiveViewSet(viewsets.ReadOnlyModelViewSet):
    queryset = AdvertisementArchive.objects.all().order_by('-created_at')
    serializer_class = AdvertisementArchiveSerializer
    permission_classes = [AuthenticatedCRUDPermission]
    filter_backends = [filters.SearchFilter, filters.OrderingFilter, DjangoFilterBackend]
    search_fields = ['Reklama_nomi','station__name']
    ordering_fields = ['created_at', 'Qurilma_narxi']
    filterset_fields = ['line','station', 'position']
    pagination_class = CustomPagination


    @action(detail=False, methods=['get'], url_path='export-excel')
    def export_excel(self, request):
        queryset = self.filter_queryset(self.get_queryset())

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Advertisement Archives"

        headers = [
            "Reklama nomi", "Qurilma turi", "Ijarachi",
            "Shartnoma raqami", "Shartnoma boshlanishi", "Shartnoma tugashi",
            "O'lchov birligi", "Qurilma narxi", "Egallagan maydon", "Shartnoma summasi",
            "Position", "Station", "Line", "Contact number", "Created at", "Created by"
        ]
        ws.append(headers)

        for ad in queryset:
            row = [
                ad.Reklama_nomi,
                str(ad.Qurilma_turi),
                str(ad.Ijarachi) if ad.Ijarachi else "",
                ad.Shartnoma_raqami,
                ad.Shartnoma_muddati_boshlanishi.strftime("%Y-%m-%d") if ad.Shartnoma_muddati_boshlanishi else "",
                ad.Shartnoma_tugashi.strftime("%Y-%m-%d") if ad.Shartnoma_tugashi else "",
                ad.O_lchov_birligi,
                float(ad.Qurilma_narxi) if ad.Qurilma_narxi else 0,
                ad.Egallagan_maydon,
                float(ad.Shartnoma_summasi) if ad.Shartnoma_summasi else 0,
                ad.position.number if ad.position else "",
                ad.position.station.name if ad.position and ad.position.station else "",
                ad.position.station.line.name if ad.position and ad.position.station and ad.position.station.line else "",
                ad.Ijarachi.contact_number if ad.Ijarachi and hasattr(ad.Ijarachi, "contact_number") else "",
                ad.created_at.strftime("%Y-%m-%d %H:%M:%S") if ad.created_at else "",
                ad.user.username if ad.user else "",
            ]
            ws.append(row)

        # ustun kengligi
        for i, column in enumerate(ws.columns, start=1):
            max_length = max((len(str(cell.value)) for cell in column if cell.value), default=0)
            ws.column_dimensions[get_column_letter(i)].width = max_length + 2

        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        response['Content-Disposition'] = 'attachment; filename=advertisement_archives.xlsx'
        wb.save(response)
        return response
    

    @action(detail=False, methods=['get'], url_path='export-pdf')
    def export_pdf(self, request):
        queryset = self.filter_queryset(self.get_queryset())

        data_list = []
        for ad in queryset:
            data_list.append({
                "image": ad.photo.url if ad.photo else "",
                "Reklama nomi": ad.Reklama_nomi,
                "Qurilma turi": ad.Qurilma_turi,
                "Ijarachi": str(ad.Ijarachi) if ad.Ijarachi else "",
                "Shartnoma raqami": ad.Shartnoma_raqami,
                "Shartnoma boshlanishi": ad.Shartnoma_muddati_boshlanishi.strftime("%Y-%m-%d") if ad.Shartnoma_muddati_boshlanishi else "",
                "Shartnoma tugashi": ad.Shartnoma_tugashi.strftime("%Y-%m-%d") if ad.Shartnoma_tugashi else "",
                "O'lchov birligi": ad.O_lchov_birligi,
                "Qurilma narxi": float(ad.Qurilma_narxi) if ad.Qurilma_narxi else 0,
                "Egallagan maydon": ad.Egallagan_maydon,
                "Shartnoma summasi": float(ad.Shartnoma_summasi) if ad.Shartnoma_summasi else 0,
                "Joyi": ad.position.number if ad.position else "",
                "Bekati": ad.position.station.name if ad.position and ad.position.station else "",
                "Liniyasi": ad.position.station.line.name if ad.position and ad.position.station and ad.position.station.line else "",
                "Bog'lanish raqami": ad.Ijarachi.contact_number if ad.Ijarachi and hasattr(ad.Ijarachi, "contact_number") else "",
                "Yaratilgan vaqti": ad.created_at.strftime("%Y-%m-%d %H:%M:%S") if ad.created_at else "",
                "kim tomonidan": ad.user.username if ad.user else "",
            })

        return generate_pdf_detail("archive_advertisements", "Advertisements Report", data_list)


class ExpiredAdvertisementViewSet(viewsets.ReadOnlyModelViewSet):
    queryset = Advertisement.objects.all()
    serializer_class = AdvertisementSerializer
    permission_classes = [permissions.IsAuthenticated]
    filter_backends = [filters.SearchFilter, filters.OrderingFilter, DjangoFilterBackend]
    search_fields = ['Reklama_nomi','position__station__name']
    ordering_fields = ['Shartnoma_tugashi', 'Shartnoma_muddati_boshlanishi']
    pagination_class = CustomPagination

    def list(self, request, *args, **kwargs):
        today = date.today()
        seven_days_later = today + timedelta(days=7)

        # Asl querysetni search va filter bilan ishlatish
        queryset = self.filter_queryset(self.get_queryset())

        expired = queryset.filter(Shartnoma_tugashi__lt=today)
        expiring_soon = queryset.filter(
            Shartnoma_tugashi__range=(today, seven_days_later)
        )

        expired_data = self.get_serializer(expired, many=True).data
        expiring_soon_data = self.get_serializer(expiring_soon, many=True).data

        return Response({
            "counts": {
                "tugagan": expired.count(),
                "haftada_tugaydigan": expiring_soon.count(),
                "umumiy": queryset.count()
            },
            "results": {
                "tugagan": expired_data,
                "haftada_tugaydigan": expiring_soon_data
            }
        })

    @action(detail=False, methods=['get'], url_path='export-expired-excel')
    def export_expired_excel(self, request):
        today = date.today()
        expired = Advertisement.objects.filter(Shartnoma_tugashi__lt=today)

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Tugagan"

        headers = [
            "Reklama nomi", "Qurilma turi", "Ijarachi",
            "Shartnoma raqami", "Shartnoma boshlanishi", "Shartnoma tugashi",
            "Position", "Station", "Status", "Created by"
        ]
        ws.append(headers)

        for ad in expired:
            ws.append([
                ad.Reklama_nomi,
                str(ad.Qurilma_turi) if ad.Qurilma_turi else "", 
                str(ad.Ijarachi) if ad.Ijarachi else "",
                ad.Shartnoma_raqami,
                ad.Shartnoma_muddati_boshlanishi.strftime("%Y-%m-%d") if ad.Shartnoma_muddati_boshlanishi else "",
                ad.Shartnoma_tugashi.strftime("%Y-%m-%d") if ad.Shartnoma_tugashi else "",
                ad.position.number if ad.position else "",
                ad.position.station.name if ad.position and ad.position.station else "",
                ad.user.username if ad.user else ""
                "tugagan"
            ])

        ws.append([])
        ws.append(["Umumiy tugagan soni:", expired.count()])

        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = "attachment; filename=expired_only.xlsx"
        wb.save(response)
        return response

    # ======================
    # Haftada tugaydiganlarni export qilish
    # ======================
    @action(detail=False, methods=['get'], url_path='export-week-excel')
    def export_week_excel(self, request):
        today = date.today()
        seven_days_later = today + timedelta(days=7)
        expiring_soon = Advertisement.objects.filter(
            Shartnoma_tugashi__range=(today, seven_days_later)
        )

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Haftada_tugaydigan"

        headers = [
            "Reklama nomi", "Qurilma turi", "Ijarachi",
            "Shartnoma raqami", "Shartnoma boshlanishi", "Shartnoma tugashi",
            "Position", "Station", "Status" ,"Created by"
        ]
        ws.append(headers)

        for ad in expiring_soon:
            ws.append([
                ad.Reklama_nomi,
                str(ad.Qurilma_turi) if ad.Qurilma_turi else "",
                str(ad.Ijarachi) if ad.Ijarachi else "",
                ad.Shartnoma_raqami,
                ad.Shartnoma_muddati_boshlanishi.strftime("%Y-%m-%d") if ad.Shartnoma_muddati_boshlanishi else "",
                ad.Shartnoma_tugashi.strftime("%Y-%m-%d") if ad.Shartnoma_tugashi else "",
                ad.position.number if ad.position else "",
                ad.position.station.name if ad.position and ad.position.station else "",
                ad.user.username if ad.user else ""
                "haftada_tugaydigan"
            ])

        ws.append([])
        ws.append(["Umumiy haftada tugaydigan soni:", expiring_soon.count()])

        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = "attachment; filename=expiring_week.xlsx"
        wb.save(response)
        return response
    


class AllAdvertisementsViewSet(viewsets.ReadOnlyModelViewSet):
    queryset = Advertisement.objects.select_related("position__station__line", "user").all()
    serializer_class = AdvertisementSerializer
    permission_classes = [permissions.IsAuthenticated]   
    filter_backends = [filters.SearchFilter, DjangoFilterBackend]
    
    # 3 ta qidiruv ustunlari
    search_fields = ["Shartnoma_raqami", "Reklama_nomi", "position__station__name"]
    filterset_fields = [ "Shartnoma_raqami", "Qurilma_turi",]

    pagination_class = CustomPagination

    @extend_schema(
        parameters=[
            OpenApiParameter(name="search", description="Qidiruv uchun matn", required=False, type=str),
            OpenApiParameter(name="Ijarachi", description="Ijarachi bo‘yicha filter", required=False, type=str),
            OpenApiParameter(name="Shartnoma_raqami", description="Shartnoma raqami bo‘yicha filter", required=False, type=str),
        ],
        responses={200: 'Excel fayl'}
    )
    @action(detail=False, methods=['get'], url_path='export-excel')
    def export_excel(self, request):
        queryset = self.filter_queryset(self.get_queryset())

        wb = Workbook()
        ws = wb.active
        ws.title = "Search Advertisements"

        headers = [
            "Reklama nomi", "Qurilma turi", "Ijarachi",
            "Shartnoma raqami", "Shartnoma boshlanishi", "Shartnoma tugashi",
            "O'lchov birligi", "Qurilma narxi", "Egallagan maydon", "Shartnoma summasi",
            "Position", "Station", "Line", "Contact number", "Created at", "Created by"
        ]
        ws.append(headers)

        for ad in queryset:
            ws.append([
                ad.Reklama_nomi,
                str(ad.Qurilma_turi),
                str(ad.Ijarachi) if ad.Ijarachi else "",
                ad.Shartnoma_raqami,
                ad.Shartnoma_muddati_boshlanishi.strftime("%Y-%m-%d") if ad.Shartnoma_muddati_boshlanishi else "",
                ad.Shartnoma_tugashi.strftime("%Y-%m-%d") if ad.Shartnoma_tugashi else "",
                ad.O_lchov_birligi,
                float(ad.Qurilma_narxi) if ad.Qurilma_narxi else 0,
                ad.Egallagan_maydon,
                float(ad.Shartnoma_summasi) if ad.Shartnoma_summasi else 0,
                ad.position.number if ad.position else "",
                ad.position.station.name if ad.position and ad.position.station else "",
                ad.position.station.line.name if ad.position and ad.position.station and ad.position.station.line else "",
                ad.Ijarachi.contact_number if ad.Ijarachi and hasattr(ad.Ijarachi, "contact_number") else "",
                ad.created_at.strftime("%Y-%m-%d %H:%M:%S") if ad.created_at else "",
                ad.user.username if ad.user else "",
            ])

        # ustun kengligi
        for i, column in enumerate(ws.columns, start=1):
            max_length = max((len(str(cell.value)) for cell in column if cell.value), default=0)
            ws.column_dimensions[get_column_letter(i)].width = max_length + 2

        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        response['Content-Disposition'] = 'attachment; filename=search_advertisements.xlsx'
        wb.save(response)
        return response
    

    
    @action(detail=False, methods=['get'], url_path='export-pdf')
    def export_pdf(self, request):
        queryset = self.filter_queryset(self.get_queryset())

        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4)
        styles = getSampleStyleSheet()
        elements = []

        elements.append(Paragraph("<b>Advertisements Report</b>", styles['Title']))
        elements.append(Spacer(1, 12))

        for ad in queryset:
            # --- Reklama rasmi ---
            if ad.photo:
                try:
                    img = Image(ad.photo.path, width=150, height=100)
                    elements.append(img)
                    elements.append(Spacer(1, 10))
                except Exception as e:
                    elements.append(Paragraph(f"<font color='red'>Rasm yuklanmadi: {str(e)}</font>", styles['Normal']))

            # --- Ma’lumotlar jadvali ---
            data = [
                ["Reklama nomi", ad.Reklama_nomi],
                ["Qurilma turi", ad.Qurilma_turi],
                ["Ijarachi", str(ad.Ijarachi) if ad.Ijarachi else ""],
                ["Shartnoma raqami", ad.Shartnoma_raqami],
                ["Shartnoma boshlanishi", ad.Shartnoma_muddati_boshlanishi.strftime("%Y-%m-%d") if ad.Shartnoma_muddati_boshlanishi else ""],
                ["Shartnoma tugashi", ad.Shartnoma_tugashi.strftime("%Y-%m-%d") if ad.Shartnoma_tugashi else ""],
                ["O'lchov birligi", ad.O_lchov_birligi],
                ["Qurilma narxi", ad.Qurilma_narxi],
                ["Egallagan maydon", ad.Egallagan_maydon],
                ["Shartnoma summasi", ad.Shartnoma_summasi],
                ["Joyi", ad.position.number if ad.position else ""],
                ["Bekati", ad.position.station.name if ad.position and ad.position.station else ""],
                ["Liniyasi", ad.position.station.line.name if ad.position and ad.position.station and ad.position.station.line else ""],
                ["Bog'lanish raqami", ad.Ijarachi.contact_number if ad.Ijarachi and hasattr(ad.Ijarachi, "contact_number") else ""],
                ["Yaratilgan vaqti", ad.created_at.strftime("%Y-%m-%d %H:%M:%S") if ad.created_at else ""],
                ["Kim tomonidan", ad.user.username if ad.user else ""],
            ]

            table = Table(data, colWidths=[150, 350])
            table.setStyle(TableStyle([
                ("GRID", (0,0), (-1,-1), 0.5, colors.grey),
                ("BACKGROUND", (0,0), (0,-1), colors.whitesmoke),
                ("VALIGN", (0,0), (-1,-1), "TOP"),
            ]))

            elements.append(table)
            elements.append(Spacer(1, 20))

        doc.build(elements)
        buffer.seek(0)

        # --- Faylni yuklab olish uchun javob ---
        return FileResponse(buffer, as_attachment=True, filename="search_advertisements.pdf")




class ShartnomaSummasiViewSet(viewsets.ModelViewSet):
    queryset = ShartnomaSummasi.objects.all().select_related('advertisement')
    serializer_class = ShartnomaSummasiSerializer
    permission_classes = [IsAuthenticated]

    def perform_create(self, serializer):
        instance = serializer.save()  # yangi to‘lov yaratildi
        ad = instance.advertisement

        # jami summani qayta hisoblash
        total_sum = ShartnomaSummasi.objects.filter(advertisement=ad).aggregate(
            total=Sum('Shartnomasummasi')
        )['total'] or 0

        ad.Shartnoma_summasi = total_sum
        ad.save()


class AdvertisementStatisticsView(APIView):
    permission_classes = [permissions.IsAuthenticated]



    @extend_schema(
        responses={
            200: AdvertisementStatisticsSerializer
        }
    )
    
    def get(self, request):
        today = now().date()
        first_day_this_month = today.replace(day=1)

        top_5_ads = Advertisement.objects.order_by("-Shartnoma_summasi")[:5]
        last_10_ads = Advertisement.objects.order_by("-created_at")[:10]


        total_count = Advertisement.objects.count()

        top_5_stations = (
            Advertisement.objects.values("position__station__name")
            .annotate(total=Count("id"))
            .order_by("-total")[:10]
        )
        hamkor_tashkilot_soni = Advertisement.objects.values("Ijarachi").distinct().count()
        reklamadan_tushadigan_umumiy_summa = Advertisement.objects.aggregate(total=Sum("Shartnoma_summasi"))["total"] or 0

        serializer = AdvertisementStatisticsSerializer({
            "top_5_ads": top_5_ads,
            "last_10_ads": last_10_ads,
            "top_5_stations": list(top_5_stations),
            "counts": [ 
                {"name": "total_count", "value": total_count, "color": "barcha reklamalar soni"}, 
                {"name": "hamkor_tashkilot_soni", "value": hamkor_tashkilot_soni, "color": "hamkor tashkilotlar soni"}, 
                {"name": "reklamadan_tushadigan_umumiy_summa", "value": reklamadan_tushadigan_umumiy_summa, "color": "reklamadan tushadigan umumiy summa"}
            ]
            
        }, context={"request": request})

        return Response(serializer.data)



class AdvertisementStatisticsViewSet(viewsets.GenericViewSet):
    permission_classes = [IsAuthenticated]
    
    def list(self, request):
        return self.all_statistics(request)

    @action(detail=False, methods=["get"], url_path="all")
    def all_statistics(self, request):
        lines_data = []

        for line in MetroLine.objects.all():
            stations_data = []
            line_ads = Advertisement.objects.filter(position__station__line=line)

            for station in Station.objects.filter(line=line):
                station_ads = Advertisement.objects.filter(position__station=station)

                # ✅ Qurilma turi bo‘yicha guruhlash
                turi_stats = (
                    station_ads.values('Qurilma_turi__qurilmaturi')
                    .annotate(count=Count('id'))
                    .order_by('-count')
                )

                turi_detail = []
                for item in turi_stats:
                    turi_nomi = item['Qurilma_turi__qurilmaturi'] or "Noma'lum"
                    turi_ads = station_ads.filter(Qurilma_turi__qurilmaturi=turi_nomi)

                    # ✅ Serializerga context qo‘shildi
                    turi_detail.append({
                        "turi": turi_nomi,
                        "reklama_soni": item['count'],
                        "reklamalar": AdvertisementSerializer(
                            turi_ads, many=True, context={"request": request}
                        ).data
                    })

                stations_data.append({
                    "bekat": station.name,
                    "umumiy_reklama_soni": station_ads.count(),
                    "qurilma_turlari": turi_detail
                })

            lines_data.append({
                "liniya": line.name,
                "umumiy_reklama_soni": line_ads.count(),
                "bekatlar": stations_data
            })

        return Response({
            "umumiy_reklama_soni": Advertisement.objects.count(),
            "liniyalar": lines_data
        })
        
    
    @action(detail=False, methods=["get"], url_path="all-pdf")
    def all_statistics_pdf(self, request):
        lines_data = []
        total_ads = 0  # jami reklamalar soni

        # 1️⃣ Collect data
        for line in MetroLine.objects.all():
            stations_data = []
            for station in Station.objects.filter(line=line):
                station_ads = Advertisement.objects.filter(position__station=station)
                total_ads += station_ads.count()  # jami reklama sonini qo'shamiz

                turi_stats = (
                    station_ads.values('Qurilma_turi__qurilmaturi')
                    .annotate(count=Count('id'))
                )

                station_dict = {"name": station.name, "turlar": {}}
                for item in turi_stats:
                    turi_nomi = item.get('Qurilma_turi__qurilmaturi') or "Noma'lum"
                    station_dict["turlar"][turi_nomi] = item['count']

                stations_data.append(station_dict)
            lines_data.append({"line": line.name, "stations": stations_data})

        if total_ads == 0:
            return Response({"detail": "PDF yaratish uchun reklama ma'lumotlari mavjud emas"}, status=400)

        # 2️⃣ Create PDF
        buffer = BytesIO()
        doc = SimpleDocTemplate(
            buffer,
            pagesize=landscape(A4),
            rightMargin=20, leftMargin=20, topMargin=20, bottomMargin=20
        )
        elements = []
        styles = getSampleStyleSheet()

        # Title
        title_style = styles['Title']
        title_style.textColor = colors.blue
        elements.append(Paragraph("Bekatlarda o'rnatilgan reklamalar statistikasi", title_style))
        elements.append(Spacer(1, 12))

        # Table header
        header_style = ParagraphStyle(name='header', textColor=colors.red, alignment=1, fontSize=10)
        all_turlari = set()
        for line in lines_data:
            for station in line["stations"]:
                all_turlari.update(station["turlar"].keys())
        all_turlari = sorted(all_turlari)
        
        header = [Paragraph("Liniya", header_style), Paragraph("Bekat", header_style)]
        for turi in all_turlari:
            header.append(Paragraph(turi, header_style))
        table_data = [header]

        # Data rows
        line_station_style = ParagraphStyle(name='line_station', alignment=1, fontSize=9, textColor=colors.darkblue)
        row_style = ParagraphStyle(name='row', alignment=1, fontSize=9, textColor=colors.black)

        for line in lines_data:
            stations = line["stations"]
            for idx, station in enumerate(stations):
                row = [
                    Paragraph(line["line"] if idx == 0 else "", line_station_style),
                    Paragraph(station["name"], line_station_style)
                ]
                for turi in all_turlari:
                    val = station["turlar"].get(turi, 0)
                    row.append(Paragraph(str(val) if val != 0 else "-", row_style))
                table_data.append(row)

        # Column widths
        page_width = landscape(A4)[0] - doc.leftMargin - doc.rightMargin
        num_turlar = len(all_turlari)
        turlar_col_width = (page_width - 220) / num_turlar if num_turlar > 0 else 50
        col_widths = [100, 120] + [turlar_col_width]*num_turlar

        table = Table(table_data, colWidths=col_widths, repeatRows=1)

        # Table style
        table_style = TableStyle([
            ('GRID', (0,0), (-1,-1), 1, colors.grey),
            ('ALIGN', (0,0), (-1,-1), 'CENTER'),
            ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
            ('TEXTCOLOR', (0,0), (-1,0), colors.red),
            ('TEXTCOLOR', (0,1), (1,-1), colors.darkblue),
            ('BACKGROUND', (0,0), (-1,0), colors.lightyellow),
            ('BACKGROUND', (0,1), (-1,-1), colors.whitesmoke),
        ])

        # Merge line cells
        current_row = 1
        for line in lines_data:
            n_stations = len(line["stations"])
            if n_stations > 1:
                table_style.add('SPAN', (0, current_row), (0, current_row + n_stations - 1))
            current_row += n_stations

        table.setStyle(table_style)
        elements.append(table)

        # --- Pastga jami reklamalar soni ---
        total_paragraph = Paragraph(
            f'<font color="red">Jami reklamalar soni:</font> <font color="darkblue">{total_ads}</font>',
            ParagraphStyle(name='total', fontSize=10)
        )
        elements.append(Spacer(1, 12))
        elements.append(total_paragraph)

        # Build PDF
        doc.build(elements)
        buffer.seek(0)
        response = HttpResponse(buffer, content_type='application/pdf')
        response['Content-Disposition'] = 'attachment; filename="metro_statistics.pdf"'
        return response

        

class IjarachiStatisticsViewSet(viewsets.GenericViewSet):
    permission_classes = [IsAuthenticated]

    def list(self, request):
        return self.all_statistics(request)


    @action(detail=False, methods=["get"], url_path="all")
    def all_statistics(self, request):
        tenants_data = []

        for tenant in Ijarachi.objects.all():
            # Shu ijarachiga tegishli barcha reklama
            tenant_ads = Advertisement.objects.filter(Ijarachi=tenant)

            # Bekatlar bo'yicha guruhlash
            stations_data = []
            # Note: positions__advertisement orqali bog‘lanish
            for station in Station.objects.filter(positions__advertisement__Ijarachi=tenant).distinct():
                station_ads = tenant_ads.filter(position__station=station)
                
                stations_data.append({
                    "bekat": station.name,
                    "reklama_soni": station_ads.count(),
                    "reklamalar": AdvertisementSerializer(
                        station_ads, many=True, context={"request": request}
                    ).data
                })

            tenants_data.append({
                "ijarachi": tenant.name,
                "umumiy_reklama_soni": tenant_ads.count(),
                "bekatlar": stations_data
            })

        return Response({
            "umumiy_ijarachi_soni": Ijarachi.objects.count(),
            "ijarachilar": tenants_data
        })      
        
        
    @action(detail=False, methods=["get"], url_path="all-pdf")
    def all_statistics_pdf(self, request):
        # 10 ta eng ko'p reklama berilgan ijarachilar
        top_tenants = (
            Ijarachi.objects.all()
            .annotate(reklama_count=Count('advertisement'))
            .order_by('-reklama_count')[:10]
        )

        if not top_tenants.exists():
            return Response(
                {"detail": "PDF yaratish uchun ijarachi ma'lumotlari mavjud emas"},
                status=400
            )

        tenant_names = [t.name for t in top_tenants]
        tenant_counts = [Advertisement.objects.filter(Ijarachi=t).count() for t in top_tenants]

        max_value = max(tenant_counts)
        step_value = max(1, int(max_value / 7))

        # PDF yaratish
        buffer = BytesIO()
        doc = SimpleDocTemplate(
            buffer,
            pagesize=landscape(A4),
            rightMargin=20,
            leftMargin=20,
            topMargin=20,
            bottomMargin=20
        )
        elements = []

        # TITLE (dark blue)
        styles = getSampleStyleSheet()
        title_style = styles["Title"]
        title_style.textColor = colors.darkblue
        elements.append(Paragraph("Top 10 Ijarachilar Reklamalari Statistikasi", title_style))
        elements.append(Spacer(1, 20))

        # CHART
        drawing = Drawing(700, 400)
        chart = HorizontalBarChart()
        chart.x = 120
        chart.y = 50
        chart.height = 300
        chart.width = 500
        chart.data = [tenant_counts]

        chart.barWidth = 22
        chart.groupSpacing = 12
        chart.barSpacing = 5

        # Shkala
        chart.valueAxis.valueMin = 0
        chart.valueAxis.valueMax = max_value + step_value
        chart.valueAxis.valueStep = step_value

        # Ijarachi nomlari
        chart.categoryAxis.categoryNames = tenant_names
        chart.categoryAxis.labels.fillColor = colors.darkblue
        chart.categoryAxis.labels.boxAnchor = 'e'

        # Bar rangi
        chart.bars[0].fillColor = colors.lightgreen

        # Bar tugagandan keyin qiymatni qo'shish (faqat 0 dan katta bo'lsa)
        for index, value in enumerate(tenant_counts):
            bar_length = value * (chart.width / (max_value + step_value))
            y_center = chart.y + index * (chart.barWidth + chart.groupSpacing) + (chart.barWidth / 2)

            if value > 0:
                value_str = f"{value:>4}"
                x_pos = chart.x + bar_length + 12
                drawing.add(
                    String(
                        x_pos,
                        y_center - 4,
                        value_str,
                        fontName='Helvetica-Bold',
                        fontSize=12,
                        fillColor=colors.black,
                        textAnchor="start"
                    )
                )

        drawing.add(chart)
        elements.append(drawing)

        # PDF yaratish
        doc.build(elements)
        buffer.seek(0)

        response = HttpResponse(buffer, content_type='application/pdf')
        response['Content-Disposition'] = 'attachment; filename="top_10_ijarachi_statistics.pdf"'
        return response









class IjarachiSumStatisticsViewSet(viewsets.GenericViewSet):
    permission_classes = [IsAuthenticated]

    def list(self, request):
        return self.all_statistics(request)

    
    @action(detail=False, methods=["get"], url_path="all")
    def all_statistics(self, request):
        tenants_data = []

        for tenant in Ijarachi.objects.all():
            tenant_ads = Advertisement.objects.filter(Ijarachi=tenant)

            # Ijarachi bo'yicha jami to'lov
            tenant_total_sum = ShartnomaSummasi.objects.filter(advertisement__in=tenant_ads).aggregate(
                total=Sum('Shartnomasummasi')
            )['total'] or 0

            # Har bir reklama uchun to'lovlar
            ads_data = []
            for ad in tenant_ads:
                ad_sum = ad.tolovlar.aggregate(total=Sum('Shartnomasummasi'))['total'] or 0
                ads_data.append({
                    "reklama": ad.Reklama_nomi,
                    "to'langan_summasi": ad_sum
                })

            # Bekatlar bo'yicha guruhlash
            stations_data = []
            for station in Station.objects.filter(positions__advertisement__Ijarachi=tenant).distinct():
                station_ads = tenant_ads.filter(position__station=station)

                # Shu bekatdagi jami to'langan summa
                station_total_sum = ShartnomaSummasi.objects.filter(advertisement__in=station_ads).aggregate(
                    total=Sum('Shartnomasummasi')
                )['total'] or 0

                station_ads_data = []
                for ad in station_ads:
                    ad_sum = ad.tolovlar.aggregate(total=Sum('Shartnomasummasi'))['total'] or 0
                    station_ads_data.append({
                        "reklama": ad.Reklama_nomi,
                        "to'langan_summasi": ad_sum
                    })

                stations_data.append({
                    "bekat": station.name,
                    "umumiy_to'langan_summasi": station_total_sum,
                    "reklamalar": station_ads_data
                })

            tenants_data.append({
                "ijarachi": tenant.name,
                "umumiy_to'langan_summasi": tenant_total_sum,
                "reklamalar": ads_data,
                "bekatlar": stations_data
            })

        return Response({
            "umumiy_ijarachi_soni": Ijarachi.objects.count(),
            "ijarachilar": tenants_data
        })
        
        
    @action(detail=False, methods=["get"], url_path="pdf")
    def all_statistics_pdf(self, request):
        year = datetime.now().year

        tenants = (
            Ijarachi.objects.annotate(
                total_sum=Sum('advertisement__tolovlar__Shartnomasummasi')
            ).order_by('-total_sum')[:10]
        )

        if not tenants.exists():
            return Response({"detail": "PDF yaratish uchun ma'lumotlar mavjud emas"}, status=400)

        h1_start, h1_end = datetime(year, 1, 1), datetime(year, 6, 30, 23, 59, 59)
        h2_start, h2_end = datetime(year, 7, 1), datetime(year, 12, 31, 23, 59, 59)

        chart_data_h1 = []
        chart_data_h2 = []
        tenant_names = []
        total_sum_all = Decimal(0)

        for tenant in tenants:
            tenant_names.append(tenant.name)
            sum_h1 = ShartnomaSummasi.objects.filter(
                advertisement__Ijarachi=tenant,
                created_at__range=(h1_start, h1_end)
            ).aggregate(total=Sum('Shartnomasummasi'))['total'] or 0

            sum_h2 = ShartnomaSummasi.objects.filter(
                advertisement__Ijarachi=tenant,
                created_at__range=(h2_start, h2_end)
            ).aggregate(total=Sum('Shartnomasummasi'))['total'] or 0

            chart_data_h1.append(float(sum_h1))
            chart_data_h2.append(float(sum_h2))
            total_sum_all += sum_h1 + sum_h2

        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=landscape(A4), rightMargin=40, leftMargin=120,
                                topMargin=20, bottomMargin=20)
        elements = []
        styles = getSampleStyleSheet()
        title_style = styles['Title']
        title_style.textColor = colors.darkblue

        def create_chart(data, color):
            drawing = Drawing(700, 400)
            chart = HorizontalBarChart()
            chart.x = 0
            chart.y = 50
            chart.height = 300
            chart.width = 600
            chart.data = [data]
            chart.barWidth = 20
            chart.groupSpacing = 10
            chart.barSpacing = 5
            chart.categoryAxis.categoryNames = tenant_names
            chart.categoryAxis.labels.boxAnchor = 'e'
            chart.categoryAxis.labels.dx = -5

            # Y-axis interval 1 mln, qiymat aniq qoladi
            interval = 1_000_000
            max_val = max(data) if data else 0
            max_val = ((max_val // interval) + 1) * interval or interval
            chart.valueAxis.valueMin = 0
            chart.valueAxis.valueMax = max_val
            chart.valueAxis.valueStep = interval
            chart.bars[0].fillColor = color

            # Qiymatlarni bar ostida ko'rsatish (an exact value)
            for i, val in enumerate(data):
                y = chart.y + i * (chart.barWidth + chart.groupSpacing)
                display_val = f"{val:,.0f}"  # aniq qiymat, yaxlitlamaymiz
                drawing.add(String(chart.width + 15, y + chart.barWidth / 2 - 4, display_val,
                           fontSize=10, fontName="Helvetica-Bold", fillColor=color))

            drawing.add(chart)
            return drawing

        # 1-bet: 1-yarim yil
        elements.append(Paragraph(f"Top 10 Ijarachilar To'lov Statistikasi — {year}-yil (1-Yarim yil)", title_style))
        elements.append(Spacer(1, 12))
        elements.append(create_chart(chart_data_h1, colors.lightblue))
        elements.append(Spacer(1, 20))

        # 2-bet: 2-yarim yil
        elements.append(PageBreak())
        elements.append(Paragraph(f"Top 10 Ijarachilar To'lov Statistikasi — {year}-yil (2-Yarim yil)", title_style))
        elements.append(Spacer(1, 12))
        elements.append(create_chart(chart_data_h2, colors.green))
        elements.append(Spacer(1, 20))

        # Jami summa pastda
        total_paragraph = Paragraph(
            f"<b>Jami to'lov summasi:</b> {total_sum_all:,.2f}",
            ParagraphStyle(name='total', fontSize=12, textColor=colors.red)
        )
        elements.append(total_paragraph)

        # --- Yangi sahifa: Bekatlar bo'yicha reklama summalari jadvali ---
        elements.append(PageBreak())
        elements.append(Paragraph(f"Bekatlar bo'yicha reklamalar va to'lov summalari — {year}-yil", title_style))
        elements.append(Spacer(1, 12))

        # Jadval sarlavhalari
        months = ['Yan', 'Fev', 'Mar', 'Apr', 'May', 'Iyun', 'Iyul', 'Avg', 'Sen', 'Okt', 'Noy', 'Dek']
        table_data = [['Bekat'] + months]   # ❗️ Jami olib tashlandi

        # Bekatlar bo‘yicha jadval
        all_stations = Station.objects.all()
        for station in all_stations:
            station_paragraph = Paragraph(
                station.name,
                ParagraphStyle(
                    name='station',
                    textColor=colors.darkblue,
                    fontName='Helvetica-Bold',
                    fontSize=10,       # ❗️ Bekatlar ko‘p bo‘lgani uchun kichikroq
                    leading=9,
                    alignment=TA_LEFT
                )
            )

            row = [station_paragraph]

            for month in range(1, 13):
                month_sum = ShartnomaSummasi.objects.filter(
                    advertisement__position__station=station,
                    created_at__year=year,
                    created_at__month=month
                ).aggregate(total=Sum('Shartnomasummasi'))['total'] or 0

                row.append(f"{month_sum:,.0f}")

            table_data.append(row)

        # Sahifa kengligi (A4 Horizontal)
        left_margin = 5
        right_margin = 5
        page_width = landscape(A4)[0] - left_margin - right_margin

        # Ustunlarni 85% qilib beramiz → 15% bo‘sh joy qoladi
        col_widths = [page_width * 0.13] + [page_width * 0.06] * 12

        table_style = TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.lightgrey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.black),
            ('ALIGN', (1, 1), (-1, -1), 'RIGHT'),
            ('ALIGN', (0, 0), (0, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ])

        # Jadvalni markazga hizalash
        t = Table(table_data, colWidths=col_widths, repeatRows=1, hAlign='CENTER')

        t.setStyle(table_style)
        elements.append(t)




        doc.build(elements)
        buffer.seek(0)
        response = HttpResponse(buffer, content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="top10_ijarachi_statistics_{year}.pdf"'
        return response
        
             

class CheckAuthView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        # Token valid bo'lsa shu API ishlaydi
        return Response({
            "detail": False
        })

    def handle_exception(self, exc):
        # Agar token noto'g'ri yoki yo'q bo'lsa
        if getattr(exc, 'status_code', None) in [401, 403]:
            return Response({"detail": True},
                            status=status.HTTP_401_UNAUTHORIZED)
        return super().handle_exception(exc)
    

class Last10AdvertisementImagesView(APIView):
    """
    Oxirgi 10 ta reklamani rasmini chiqaradigan API
    """
    permission_classes = [permissions.IsAuthenticated]  

    def get(self, request):
        ads = Advertisement.objects.order_by("-created_at")[:5]
        data = [
            {
                "id": ad.id,
                "name": ad.Reklama_nomi,
                "photo": request.build_absolute_uri(ad.photo.url) if ad.photo else None
            }
            for ad in ads
        ]
        return Response(data)
    


class StatisticsCountAPI(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        data = {
            "metro_lines": MetroLine.objects.count(),
            "stations": Station.objects.count(),
            "positions": Position.objects.count(),
            "advertisements": Advertisement.objects.count(),
            "advertisement_archives": AdvertisementArchive.objects.count(),

        }
        return Response(data)
    

# Harakat tarkiblari uchun reklamalr ======================================================================================================================


class DepoViewSet(viewsets.ModelViewSet):
    queryset = Depo.objects.all()
    serializer_class = DepoSerializer
    permission_classes = [AuthenticatedCRUDPermission]
    pagination_class = CustomPagination

class TarkibViewSet(viewsets.ModelViewSet):
    queryset = HarakatTarkibi.objects.all()
    serializer_class = HarakatTarkibiSerializer
    permission_classes = [AuthenticatedCRUDPermission]
    filter_backends = [DjangoFilterBackend, filters.SearchFilter]
    filterset_fields = ['depo']
    search_fields = ['tarkib', 'depo__nomi']
    pagination_class = CustomPagination
    
    
class Tarkibimage(APIView):
    permission_classes = [IsAuthenticated]

    def put(self, request, pk):
        try:
            tarkib = HarakatTarkibi.objects.get(pk=pk)
        except HarakatTarkibi.DoesNotExist:
            return Response({"error": "Bunday harakat tarkibi topilmadi"}, status=404)

        image = request.FILES.get("schema_image")
        if not image:
            return Response({"error": "Rasm yuborilmadi"}, status=400)

        tarkib.schema_image = image
        tarkib.save()
        return Response({"message": "Rasm yangilandi", "id": tarkib.id})




class TarkibPositionViewSet(viewsets.ModelViewSet):
    queryset = (
        TarkibPosition.objects
        .select_related('harakat_tarkibi')
        .prefetch_related('tarkib_advertisement')  
        .all()
        .order_by('-id')
    )
    serializer_class = TarkibPositionSerializer
    permission_classes = [AuthenticatedCRUDPermission]
    filter_backends = [DjangoFilterBackend, filters.OrderingFilter, filters.SearchFilter]

    filterset_fields = ['harakat_tarkibi']   # advertisement emas!
    search_fields = ['position', 'harakat_tarkibi__tarkib']
    pagination_class = CustomPagination

    def perform_create(self, serializer):
        serializer.save(
            created_by=self.request.user,
            created_at=timezone.now()
        )

    def perform_destroy(self, instance):
       
        # biriktirilgan tarkib reklamalari
        ads = getattr(instance, "tarkib_advertisement", None)

        # agar mavjud bo‘lsa
        if ads:
            for ad in ads.all():

                TarkibAdvertisementArchive.objects.create(
                    original_ad=ad,
                    user=self.request.user,
                    tarkib_position=instance,
                    harakat_tarkibi=instance.harakat_tarkibi,

                    Reklama_nomi=ad.Reklama_nomi,
                    Qurilma_turi=ad.Qurilma_turi,
                    Ijarachi=ad.Ijarachi,
                    Shartnoma_raqami=ad.Shartnoma_raqami,
                    Shartnoma_muddati_boshlanishi=ad.Shartnoma_muddati_boshlanishi,
                    Shartnoma_tugashi=ad.Shartnoma_tugashi,
                    O_lchov_birligi=ad.O_lchov_birligi,
                    Qurilma_narxi=ad.Qurilma_narxi,
                    Egallagan_maydon=ad.Egallagan_maydon,
                    Shartnoma_summasi=ad.Shartnoma_summasi,
                    Shartnoma_fayl=ad.Shartnoma_fayl,
                    photo=ad.photo,
                    contact_number=ad.contact_number,
                )

                ad.delete()

        instance.delete()



class TarkibShardnomaSummasiViewSet(viewsets.ModelViewSet):
    queryset = TarkibShartnomaSummasi.objects.all().select_related('reklama')
    serializer_class = TarkibShartnomaSummasiSerializer
    permission_classes = [IsAuthenticated]

    def perform_create(self, serializer):
        instance = serializer.save()  # yangi to‘lov yaratildi
        ad = instance.reklama

        # jami summani qayta hisoblash
        total_sum = TarkibShartnomaSummasi.objects.filter(reklama=ad).aggregate(
            total=Sum('Shartnomasummasi')
        )['total'] or 0

        ad.Shartnoma_summasi = total_sum
        ad.save()

class TarkibAdvertisementViewSet(viewsets.ModelViewSet):
    queryset = TarkibAdvertisement.objects.select_related(
        'position', 'position__harakat_tarkibi', 'user', 'Ijarachi'
    ).prefetch_related('tarkibtolovlar').all().order_by("-id")

    permission_classes = [AuthenticatedCRUDPermission]
    filter_backends = [filters.SearchFilter, filters.OrderingFilter, DjangoFilterBackend]
    search_fields = ['Reklama_nomi', 'Shartnoma_raqami']
    ordering_fields = ['Qurilma_narxi']
    filterset_fields = ['position', 'Ijarachi', 'position__harakat_tarkibi']
    serializer_class = TarkibAdvertisementSerializer

    # ---- SERIALIZERLARNI ACTIONGA KO‘RA TANLASH ----
    def get_serializer_class(self):
        if self.action == 'create':
            return CreateTarkibAdvertisementSerializer
        elif self.action in ['update', 'partial_update']:
            return UpdateTarkibAdvertisementSerializer
        return TarkibAdvertisementSerializer

    # ---- CREATE ----
    def perform_create(self, serializer):
        serializer.save(user=self.request.user)

    # ---- UPDATE ----
    def perform_update(self, serializer):
        old_ad = serializer.instance

        # ---- ARXIVGA ESKISI YOZILADI ----
        archive = TarkibAdvertisementArchive.objects.create(
            original_ad=old_ad,
            user=self.request.user,
            position=old_ad.position,
            tarkib=old_ad.position.harakat_tarkibi if old_ad.position else None,
            Reklama_nomi=old_ad.Reklama_nomi,
            Qurilma_turi=old_ad.Qurilma_turi,
            Ijarachi=old_ad.Ijarachi,
            Shartnoma_raqami=old_ad.Shartnoma_raqami,
            Shartnoma_muddati_boshlanishi=old_ad.Shartnoma_muddati_boshlanishi,
            Shartnoma_tugashi=old_ad.Shartnoma_tugashi,
            O_lchov_birligi=old_ad.O_lchov_birligi,
            Qurilma_narxi=old_ad.Qurilma_narxi,
            Egallagan_maydon=old_ad.Egallagan_maydon,
            Shartnoma_summasi=old_ad.Shartnoma_summasi,
            Shartnoma_fayl=old_ad.Shartnoma_fayl,
            photo=old_ad.photo,
        )

        # ---- To‘lovlarni arxivga o‘tkazish ----
        for t in old_ad.tarkibtolovlar.all():
            TarkibAdvertisementArchiveShartnomaSummasi.objects.create(
                reklama_archive=archive,
                Shartnomasummasi=t.Shartnomasummasi,
                comment=t.comment,
                created_at=t.created_at
            )

        # ---- Reklamani yangilash ----
        updated = serializer.save()
        return updated

    def update(self, request, *args, **kwargs):
        partial = kwargs.pop('partial', False)
        instance = self.get_object()
        serializer = self.get_serializer(instance, data=request.data, partial=partial)
        serializer.is_valid(raise_exception=True)
        updated_instance = self.perform_update(serializer)
        return Response(self.get_serializer(updated_instance).data)

    # ---- DELETE ----
    def perform_destroy(self, instance):
        tarkib = instance.position.harakat_tarkibi if instance.position else None

        archive = TarkibAdvertisementArchive.objects.create(
            original_ad=instance,
            user=self.request.user,
            position=instance.position,
            tarkib=tarkib,
            Reklama_nomi=instance.Reklama_nomi,
            Qurilma_turi=instance.Qurilma_turi,
            Ijarachi=instance.Ijarachi,
            Shartnoma_raqami=instance.Shartnoma_raqami,
            Shartnoma_muddati_boshlanishi=instance.Shartnoma_muddati_boshlanishi,
            Shartnoma_tugashi=instance.Shartnoma_tugashi,
            O_lchov_birligi=instance.O_lchov_birligi,
            Qurilma_narxi=instance.Qurilma_narxi,
            Egallagan_maydon=instance.Egallagan_maydon,
            Shartnoma_summasi=instance.Shartnoma_summasi,
            Shartnoma_fayl=instance.Shartnoma_fayl,
            photo=instance.photo,
        )

        for t in instance.tarkibtolovlar.all():
            TarkibAdvertisementArchiveShartnomaSummasi.objects.create(
                reklama_archive=archive,
                Shartnomasummasi=t.Shartnomasummasi,
                comment=t.comment,
                created_at=t.created_at
            )

        instance.delete()




class TarkibAdvertisementArchiveViewSet(viewsets.ReadOnlyModelViewSet):
    queryset = TarkibAdvertisementArchive.objects.select_related(
        "position", "position__harakat_tarkibi", "Ijarachi", "user"
    ).prefetch_related("tarkibtolovlar").all().order_by("-id")

    serializer_class = TarkibAdvertisementArchiveSerializer
    permission_classes = [AuthenticatedCRUDPermission]
    filter_backends = [filters.SearchFilter, filters.OrderingFilter, DjangoFilterBackend]

    # ---- search & filter ----
    search_fields = ['Reklama_nomi', 'position__harakat_tarkibi__name']
    ordering_fields = ['Qurilma_narxi']
    filterset_fields = ['position', 'tarkib', 'Ijarachi']  # harakat_tarkibi o‘rniga 'tarkib'
    pagination_class = CustomPagination


    # =====================
    #   EXPORT EXCEL
    # =====================
    @action(detail=False, methods=['get'], url_path='export-excel')
    def export_excel(self, request):
        queryset = self.filter_queryset(self.get_queryset())

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Tarkib Archive"

        headers = [
            "Reklama nomi", "Qurilma turi", "Ijarachi",
            "Shartnoma raqami", "Boshlanish", "Tugash",
            "Narx", "Maydon", "Jami to‘lov", "Tarkib",
            "Joy raqami", "Yaratgan", "Sanasi"
        ]
        ws.append(headers)

        for ad in queryset:
            jami_tolov = sum(t.Shartnomasummasi for t in ad.tolovlar.all())

            ws.append([
                ad.Reklama_nomi,
                str(ad.Qurilma_turi),
                str(ad.Ijarachi.name if ad.Ijarachi else ""),
                ad.Shartnoma_raqami,
                ad.Shartnoma_muddati_boshlanishi,
                ad.Shartnoma_tugashi,
                float(ad.Qurilma_narxi) if ad.Qurilma_narxi else 0,
                ad.Egallagan_maydon,
                float(jami_tolov),
                ad.position.harakat_tarkibi.name if ad.position else "",
                ad.position.number if ad.position else "",
                ad.user.username if ad.user else "",
                ad.created_at.strftime("%Y-%m-%d %H:%M:%S"),
            ])

        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename=tarkib_archive.xlsx'
        wb.save(response)
        return response

    # =====================
    #   EXPORT PDF
    # =====================
    @action(detail=False, methods=['get'], url_path='export-pdf')
    def export_pdf(self, request):
        queryset = self.filter_queryset(self.get_queryset())

        data_list = []

        for ad in queryset:
            data_list.append({
                "image": ad.photo.url if ad.photo else "",
                "Reklama nomi": ad.Reklama_nomi,
                "Qurilma turi": ad.Qurilma_turi,
                "Ijarachi": ad.Ijarachi.name if ad.Ijarachi else "",
                "Shartnoma raqami": ad.Shartnoma_raqami,
                "Boshlanish": ad.Shartnoma_muddati_boshlanishi,
                "Tugash": ad.Shartnoma_tugashi,
                "Narx": float(ad.Qurilma_narxi) if ad.Qurilma_narxi else 0,
                "Maydon": ad.Egallagan_maydon,
                "Tarkib": ad.position.harakat_tarkibi.name if ad.position else "",
                "Joy": ad.position.number if ad.position else "",
                "Sanasi": ad.created_at.strftime("%Y-%m-%d %H:%M:%S"),
                "Kim tomonidan": ad.user.username if ad.user else "",
            })

        return generate_pdf_detail("tarkib_archive", "Tarkib Archive", data_list)



class TarkibExpiredAdvertisementViewSet(viewsets.ReadOnlyModelViewSet):
    queryset = TarkibAdvertisement.objects.select_related(
        "position", "position__harakat_tarkibi", "user"
    ).all()

    serializer_class = TarkibAdvertisementSerializer
    permission_classes = [permissions.IsAuthenticated]
    filter_backends = [filters.SearchFilter, filters.OrderingFilter, DjangoFilterBackend]

    search_fields = ['Reklama_nomi', 'position__harakat_tarkibi__name']
    ordering_fields = ['Shartnoma_tugashi', 'Shartnoma_muddati_boshlanishi']
    pagination_class = CustomPagination

    def list(self, request, *args, **kwargs):
        today = date.today()
        week = today + timedelta(days=7)

        queryset = self.filter_queryset(self.get_queryset())

        expired = queryset.filter(Shartnoma_tugashi__lt=today)
        expiring_soon = queryset.filter(Shartnoma_tugashi__range=(today, week))

        return Response({
            "counts": {
                "tugagan": expired.count(),
                "haftada_tugaydigan": expiring_soon.count(),
                "umumiy": queryset.count(),
            },
            "results": {
                "tugagan": self.get_serializer(expired, many=True).data,
                "haftada_tugaydigan": self.get_serializer(expiring_soon, many=True).data,
            }
        })




class TarkibAllAdvertisementViewSet(viewsets.ReadOnlyModelViewSet):
    """
    Read-only viewset for all Tarkib Advertisements (Depo/Harakat tarkibi based)
    """
    queryset = TarkibAdvertisement.objects.select_related(
        'position', 'position__harakat_tarkibi', 'position__harakat_tarkibi__depo',
        'Ijarachi', 'user', 'Qurilma_turi'
    ).prefetch_related('tarkibtolovlar').all().order_by('-id')

    serializer_class = TarkibAdvertisementSerializer
    permission_classes = [permissions.IsAuthenticated]
    filter_backends = [filters.SearchFilter, filters.OrderingFilter, DjangoFilterBackend]
    search_fields = ['Reklama_nomi', 'Shartnoma_raqami']
    ordering_fields = ['Qurilma_narxi']
    filterset_fields = ['position', 'Ijarachi', 'position__harakat_tarkibi', 'position__harakat_tarkibi__depo']

    # ===== EXCEL EXPORT =====
    @extend_schema(
        parameters=[
            OpenApiParameter(name="search", description="Qidiruv matni", required=False, type=str),
            OpenApiParameter(name="Ijarachi", description="Ijarachi bo‘yicha filter", required=False, type=str),
            OpenApiParameter(name="Shartnoma_raqami", description="Shartnoma raqami bo‘yicha filter", required=False, type=str),
        ],
        responses={200: 'Excel fayl'}
    )
    @action(detail=False, methods=['get'], url_path='export-excel')
    def export_excel(self, request):
        queryset = self.filter_queryset(self.get_queryset())

        wb = Workbook()
        ws = wb.active
        ws.title = "Tarkib Advertisements"

        headers = [
            "Reklama nomi", "Qurilma turi", "Ijarachi",
            "Shartnoma raqami", "Shartnoma boshlanishi", "Shartnoma tugashi",
            "O'lchov birligi", "Qurilma narxi", "Egallagan maydon", "Shartnoma summasi",
            "Position", "Harakat tarkibi", "Depo", "Contact number", "Created at", "Created by", "Jami to'lov"
        ]
        ws.append(headers)

        for ad in queryset:
            ws.append([
                ad.Reklama_nomi,
                str(ad.Qurilma_turi) if ad.Qurilma_turi else "",
                str(ad.Ijarachi) if ad.Ijarachi else "",
                ad.Shartnoma_raqami,
                ad.Shartnoma_muddati_boshlanishi.strftime("%d-%m-%Y") if ad.Shartnoma_muddati_boshlanishi else "",
                ad.Shartnoma_tugashi.strftime("%d-%m-%Y") if ad.Shartnoma_tugashi else "",
                ad.O_lchov_birligi,
                float(ad.Qurilma_narxi) if ad.Qurilma_narxi else 0,
                ad.Egallagan_maydon,
                float(ad.Shartnoma_summasi) if ad.Shartnoma_summasi else 0,
                ad.position.position if ad.position else "",
                ad.position.harakat_tarkibi.tarkib if ad.position and ad.position.harakat_tarkibi else "",
                ad.position.harakat_tarkibi.depo.nomi if ad.position and ad.position.harakat_tarkibi and ad.position.harakat_tarkibi.depo else "",
                ad.Ijarachi.contact_number if ad.Ijarachi and hasattr(ad.Ijarachi, "contact_number") else "",
                ad.created_at.strftime("%d-%m-%Y %H:%M:%S") if ad.created_at else "",
                ad.user.username if ad.user else "",
                float(sum(t.Shartnomasummasi for t in ad.tarkibtolovlar.all()))
            ])

        for i, column in enumerate(ws.columns, start=1):
            max_length = max((len(str(cell.value)) for cell in column if cell.value), default=0)
            ws.column_dimensions[get_column_letter(i)].width = max_length + 2

        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        response['Content-Disposition'] = 'attachment; filename=tarkib_all_advertisements.xlsx'
        wb.save(response)
        return response

    # ===== PDF EXPORT =====
    @action(detail=False, methods=['get'], url_path='export-pdf')
    def export_pdf(self, request):
        queryset = self.filter_queryset(self.get_queryset())
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4)
        styles = getSampleStyleSheet()
        elements = []

        elements.append(Paragraph("<b>Tarkib Advertisements Report</b>", styles['Title']))
        elements.append(Spacer(1, 12))

        for ad in queryset:
            data = [
                ["Reklama nomi", ad.Reklama_nomi],
                ["Qurilma turi", str(ad.Qurilma_turi) if ad.Qurilma_turi else ""],
                ["Ijarachi", str(ad.Ijarachi) if ad.Ijarachi else ""],
                ["Shartnoma raqami", ad.Shartnoma_raqami],
                ["Shartnoma boshlanishi", ad.Shartnoma_muddati_boshlanishi.strftime("%d-%m-%Y") if ad.Shartnoma_muddati_boshlanishi else ""],
                ["Shartnoma tugashi", ad.Shartnoma_tugashi.strftime("%d-%m-%Y") if ad.Shartnoma_tugashi else ""],
                ["O'lchov birligi", ad.O_lchov_birligi],
                ["Qurilma narxi", ad.Qurilma_narxi],
                ["Egallagan maydon", ad.Egallagan_maydon],
                ["Shartnoma summasi", ad.Shartnoma_summasi],
                ["Position", ad.position.position if ad.position else ""],
                ["Harakat tarkibi", ad.position.harakat_tarkibi.tarkib if ad.position and ad.position.harakat_tarkibi else ""],
                ["Depo", ad.position.harakat_tarkibi.depo.nomi if ad.position and ad.position.harakat_tarkibi and ad.position.harakat_tarkibi.depo else ""],
                ["Contact number", ad.Ijarachi.contact_number if ad.Ijarachi and hasattr(ad.Ijarachi, "contact_number") else ""],
                ["Created at", ad.created_at.strftime("%d-%m-%Y %H:%M:%S") if ad.created_at else ""],
                ["Created by", ad.user.username if ad.user else ""],
                ["Jami to'lov", float(sum(t.Shartnomasummasi for t in ad.tarkibtolovlar.all()))]
            ]
            table = Table(data, colWidths=[150, 350])
            table.setStyle(TableStyle([
                ("GRID", (0,0), (-1,-1), 0.5, colors.grey),
                ("BACKGROUND", (0,0), (0,-1), colors.whitesmoke),
                ("VALIGN", (0,0), (-1,-1), "TOP"),
            ]))
            elements.append(table)
            elements.append(Spacer(1, 20))

        doc.build(elements)
        buffer.seek(0)
        return FileResponse(buffer, as_attachment=True, filename="tarkib_all_advertisements.pdf")